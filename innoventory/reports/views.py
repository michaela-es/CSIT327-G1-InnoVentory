from django.shortcuts import render
from django.db.models import Sum
from django.utils.dateparse import parse_date
from sales.models import Sale
from products.models import Product, Category
from stocks.models import Stocks
from datetime import datetime
from django.db.models import Sum, Count
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def export_excel(request):
    # Re-run the filtered query to get fresh data
    start_date = _parse_date_or_none(request.GET.get('start_date'))
    end_date = _parse_date_or_none(request.GET.get('end_date'))
    product_q = request.GET.get('product', '').strip()
    category_q = request.GET.get('category', '').strip()

    sales_qs = Sale.objects.select_related('product_sold')

    if start_date:
        sales_qs = sales_qs.filter(sales_date__date__gte=start_date)
    if end_date:
        sales_qs = sales_qs.filter(sales_date__date__lte=end_date)
    if product_q:
        sales_qs = sales_qs.filter(product_sold__name__icontains=product_q)
    if category_q:
        sales_qs = sales_qs.filter(product_sold__category__name__icontains=category_q)

    # Get date list and prepare summaries
    date_list = (
        sales_qs
        .values_list('sales_date__date', flat=True)
        .distinct()
        .order_by('sales_date__date')
    )

    summaries = []
    for d in date_list:
        day_qs = sales_qs.filter(sales_date__date=d)
        total_sales = day_qs.count()
        total_revenue = day_qs.aggregate(total=Sum('total'))['total'] or 0

        top = (
            day_qs.values('product_sold__name')
            .annotate(qty=Sum('product_qty'))
            .order_by('-qty')
            .first()
        )
        top_product = top['product_sold__name'] if top else ''

        summaries.append({
            'date': d,
            'total_sales': total_sales,
            'total_revenue': total_revenue,
            'top_product': top_product,
        })

    # Calculate totals
    grand_total_sales = sales_qs.count()
    grand_total_revenue = sales_qs.aggregate(total=Sum('total'))['total'] or 0

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Summary"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Date
    ws.column_dimensions['B'].width = 15  # Total Sales
    ws.column_dimensions['C'].width = 20  # Total Revenue
    ws.column_dimensions['D'].width = 30  # Top Product

    # Write header with styling
    headers = ["Date", "Total Sales", "Total Revenue", "Top Product"]
    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Populate rows
    for s in summaries:
        ws.append([
            s['date'].strftime('%B %d, %Y'),
            f"{s['total_sales']} sales",
            f"₱{s['total_revenue']:,.2f}",
            s['top_product']
        ])

    # Add totals row with styling
    total_row = [
        "TOTAL",
        f"{grand_total_sales} sales",
        f"₱{grand_total_revenue:,.2f}",
        "—"
    ]
    ws.append(total_row)
    last_row = ws[ws.max_row]
    for cell in last_row:
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal='center')

    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "sales_report"
    if start_date and end_date:
        filename += f"_{start_date:%Y%m%d}-{end_date:%Y%m%d}"
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    wb.save(response)
    return response

def _parse_date_or_none(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return None


def report_dashboard(request):
    # Filters
    start_date = _parse_date_or_none(request.GET.get('start_date'))
    end_date = _parse_date_or_none(request.GET.get('end_date'))
    product_q = request.GET.get('product', '').strip()
    category_q = request.GET.get('category', '').strip()

    sales_qs = Sale.objects.select_related('product_sold')

    if start_date:
        sales_qs = sales_qs.filter(sales_date__date__gte=start_date)
    if end_date:
        sales_qs = sales_qs.filter(sales_date__date__lte=end_date)
    if product_q:
        sales_qs = sales_qs.filter(product_sold__name__icontains=product_q)
    if category_q:
        sales_qs = sales_qs.filter(product_sold__category__name__icontains=category_q)

    # Aggregate per date summary
    date_list = (
        sales_qs
        .values_list('sales_date__date', flat=True)
        .distinct()
        .order_by('sales_date__date')
    )

    summaries = []
    for d in date_list:
        day_qs = sales_qs.filter(sales_date__date=d)
        total_sales = day_qs.count()
        total_revenue = day_qs.aggregate(total=Sum('total'))['total'] or 0

        # top product of the day
        top = (
            day_qs.values('product_sold__name')
            .annotate(qty=Sum('product_qty'))
            .order_by('-qty')
            .first()
        )
        top_product = top['product_sold__name'] if top else ''

        summaries.append({
            'date': d,
            'total_sales': total_sales,
            'total_revenue': total_revenue,
            'top_product': top_product,
        })

    # ✅ Compute totals across all filtered sales
    grand_total_sales = sales_qs.count()
    grand_total_revenue = sales_qs.aggregate(total=Sum('total'))['total'] or 0

    top_overall = (
        sales_qs.values('product_sold__name')
        .annotate(qty=Sum('product_qty'))
        .order_by('-qty')
        .first()
    )
    grand_top_product = top_overall['product_sold__name'] if top_overall else ''

    totals = {
        "total_sales": grand_total_sales,
        "total_revenue": grand_total_revenue,
        "top_product": grand_top_product,
    }

    # Provide choices for filters
    products = Product.objects.order_by('name')[:200]
    categories = Category.objects.order_by('name')

    context = {
        'summaries': summaries,
        'totals': totals,    # ✅ Added totals
        'products': products,
        'categories': categories,
        'filters': {
            'start_date': start_date,
            'end_date': end_date,
            'product': product_q,
            'category': category_q,
        }
    }

    return render(request, 'reports/report_dashboard.html', context)
