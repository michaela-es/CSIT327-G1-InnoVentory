# utils.py
import pandas as pd
from django.db import transaction
from .models import Product, Category
from suppliers.models import Supplier

def import_products_from_excel(file):
    default_supplier, _ = Supplier.objects.get_or_create(
        name='Unknown Supplier',
        defaults={'contact': 'Not provided', 'address': 'Not provided'}
    )
    default_category, _ = Category.objects.get_or_create(name='Default Category')

    df = pd.read_excel(file)
    df.columns = [c.strip() for c in df.columns]

    required_columns = ['name', 'price', 'stock_quantity']
    if not all(col in df.columns for col in required_columns):
        raise ValueError(f"Excel file must contain columns: {', '.join(required_columns)}")

    created = 0
    updated = 0
    skipped = 0
    errors = []

    existing_products = Product.objects.all()
    existing_map = {p.name.lower(): p for p in existing_products}

    products_to_create = []
    products_to_update = []

    for index, row in df.iterrows():
        try:
            name = str(row['name']).strip()
            if not name:
                skipped += 1
                continue

            price = float(row['price'])
            stock_quantity = int(row['stock_quantity'])
            category_name = str(row.get('category', '')).strip()
            supplier_name = str(row.get('supplier', '')).strip()

            category = Category.objects.get_or_create(
                name=category_name or default_category.name
            )[0]
            supplier = Supplier.objects.get_or_create(
                name=supplier_name or default_supplier.name,
                defaults={'contact': 'Not provided', 'address': 'Not provided'}
            )[0]

            key = name.lower()
            if key in existing_map:
                product = existing_map[key]
                product.stock_quantity += stock_quantity
                product.price = price
                product.category = category
                product.supplier = supplier
                products_to_update.append(product)
            else:
                product = Product(
                    name=name,
                    price=price,
                    stock_quantity=stock_quantity,
                    category=default_category,
                    supplier=supplier
                )
                products_to_create.append(product)
                existing_map[key] = product

        except Exception as e:
            errors.append({'row': index + 2, 'error': str(e)})
            skipped += 1

    with transaction.atomic():
        if products_to_create:
            Product.objects.bulk_create(products_to_create)
            created += len(products_to_create)
        if products_to_update:
            Product.objects.bulk_update(
                products_to_update,
                ['stock_quantity', 'price', 'category', 'supplier']
            )
            updated += len(products_to_update)

    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors,
        'total': created + updated + skipped
    }


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from .models import Product


def generate_low_stock_excel(products, filename="low_stock_products.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Low Stock Products"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")

    ws.column_dimensions['A'].width = 40  # Product Name
    ws.column_dimensions['B'].width = 30  # Supplier Name
    ws.column_dimensions['C'].width = 30  # Supplier Contact Info

    headers = ["Product Name", "Supplier", "Contact Info"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for p in products:
        ws.append([
            p.name,
            p.supplier.name if p.supplier else "",
            p.supplier.contact if p.supplier else ""
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
